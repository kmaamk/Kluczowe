from django.contrib import admin
from django.forms import Textarea
from klucz import models
from .models import Question
import csv
from django.utils.encoding import smart_str
import xlwt
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone


class KluczAdmin(admin.ModelAdmin):


    def get_form(self, request, obj=None, **kwargs):
        form = super(KluczAdmin, self).get_form(request, obj, **kwargs)
        form.base_fields['choice_text'].widget.attrs['style'] = 'width: 65em; height:5em; '

        return form


    def export_csv(modeladmin, request, queryset):

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=kluczowe.csv'
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow([
            smart_str(u"Id"),
            smart_str(u"Odd"),
            smart_str(u"Data_wysyłki"),
            smart_str(u"Kluczowe"),
        ])
        for obj in queryset:
            writer.writerow([
                smart_str(obj.pk),
                smart_str(obj.branch),
                smart_str(obj.created_date),
                smart_str(obj.choice_text),

            ])
        return response
    export_csv.short_description = u"Export CSV"

    def export_xls(modeladmin, request, queryset):

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=kluczowe.xls'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("MyModel")

        row_num = 0

        columns = [
            (u"Id", 2000),
            (u"Odd", 2000),
            (u"Data wysyłki", 8000),
            (u"Kluczowe", 16000),
        ]

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            # set column width
            ws.col(col_num).width = columns[col_num][1]

        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1

        for obj in queryset:

            row_num += 1
            row = [
                obj.pk,
                obj.branch,
                obj.created_date,
                obj.choice_text,
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response

    export_xls.short_description = u"Export XLS"

    def export_xlsx(modeladmin, request, queryset):

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "MyModel"

        row_num = 0

        columns = [
            (u"Id", 15),
            (u"Odd", 15),
            (u"Data wysyłki", 20),
            (u"Kluczowe", 70),
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            #c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in queryset:
            row_num += 1
            row = [
                obj.pk,
                obj.branch,
                obj.created_date,
                obj.choice_text,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                #c.style.alignment.wrap_text = True

        wb.save(response)
        return response

    export_xlsx.short_description = u"Export XLSX"

    actions = [export_csv, export_xlsx]

admin.site.register(Question, KluczAdmin)

#textwrap.wrap('choice_text', width=70, **kwargs)
# Register your models here.
